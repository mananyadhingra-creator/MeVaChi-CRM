from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import or_
from datetime import datetime, date


def export_to_excel(

    query,

    headers,

    fields,

    filename

):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Data"

    # ----------------------------
    # Header Row
    # ----------------------------

    for column, header in enumerate(

        headers,

        start=1

    ):

        cell = sheet.cell(

            row=1,

            column=column,

            value=header

        )

        cell.font = Font(

            bold=True

        )

    # ----------------------------
    # Data Rows
    # ----------------------------

    row_number = 2

    for record in query.all():

        for column, field in enumerate(

            fields,

            start=1

        ):

            value = getattr(record, field, "")

            if value is None:
                value = ""

            elif isinstance(value, (datetime, date)):
                value = value.strftime("%d-%m-%Y")


            sheet.cell(

                row=row_number,

                column=column,

                value=value

            )

        row_number += 1

    output = BytesIO()

    workbook.save(

        output

    )

    output.seek(0)

    return send_file(

        output,

        download_name=filename,

        as_attachment=True,

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

def apply_export_filters(

    query,

    model,

    request,

    date_field=None,

    status_field=None,

    owner_field=None,

    search_fields=None,

    custom_filters=None

):

    # ======================================
    # SEARCH
    # ======================================

    search = request.values.get(

        'search'

    )

    if search and search_fields:

        conditions = []

        for field in search_fields:

            conditions.append(

                getattr(

                    model,

                    field

                ).ilike(

                    f'%{search}%'

                )

            )

        query = query.filter(

            or_(

                *conditions

            )

        )

    # ======================================
    # DATE RANGE
    # ======================================

    start_date = request.values.get(

        'start_date'

    )

    end_date = request.values.get(

        'end_date'

    )

    if date_field:

        if start_date:

            query = query.filter(

                getattr(

                    model,

                    date_field

                ) >= datetime.strptime(

                    start_date,

                    '%Y-%m-%d'

                ).date()

            )

        if end_date:

            query = query.filter(

                getattr(

                    model,

                    date_field

                ) <= datetime.strptime(

                    end_date,

                    '%Y-%m-%d'

                ).date()

            )

    # ======================================
    # STATUS
    # ======================================

    status = request.values.get(

        'status'

    )

    if status and status_field:

        query = query.filter(

            getattr(

                model,

                status_field

            ) == status

        )

    # ======================================
    # RECORD OWNER
    # ======================================

    owner = request.values.get(

        'record_owner'

    )

    if owner and owner_field:

        query = query.filter(

            getattr(

                model,

                owner_field

            ) == owner

        )

    # ======================================
    # CUSTOM FILTERS
    # ======================================

    if custom_filters:

        for form_field, model_field in custom_filters.items():

            value = request.values.get(

                form_field

            )

            if value:

                query = query.filter(

                    getattr(

                        model,

                        model_field

                    ) == value

                )

    return query

def export_filtered_query(

    query,

    headers,

    fields,

    filename

):

    return export_to_excel(

        query=query,

        headers=headers,

        fields=fields,

        filename=filename

    )

